import asyncio
import copy
import math
import re
import zipfile
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from pathlib import PurePosixPath

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.errors import ValidationError

OFFICE_RATE = Decimal("0.03")
RESULT_HEADERS = ("السعر بالأوقية", "نسبة المكتب", "الشحن", "سعر الطياح")
ALIASES = {
    "PRICE": {"PRICE", "UNIT PRICE", "RMB PRICE", "PRICE RMB", "UNIT PRICE RMB"},
    "PCS": {"PCS", "PCS CTN", "PCS PER CTN", "QTY CTN", "PIECES CTN"},
    "CBM": {"CBM", "CBM CTN", "CARTON CBM", "CBM PER CTN"},
}


def normalize_header(value):
    text = str(value if value is not None else "").replace("\xa0", " ")
    text = re.sub(r"[\r\n]+", " ", text.strip().upper())
    text = re.sub(r"[^\w\u0600-\u06ff]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


class ExcelPricingService:
    def __init__(self, settings):
        self.settings = settings

    async def transform(self, source_bytes, rmb_rate, shipping_cost_per_cbm):
        rate = self._parameter(rmb_rate, "معامل تحويل RMB غير صالح.")
        shipping_cost = self._parameter(shipping_cost_per_cbm, "تكلفة الشحن لكل CBM غير صالحة.")
        self._validate_archive(source_bytes)
        return await asyncio.to_thread(self._transform_sync, source_bytes, rate, shipping_cost)

    @staticmethod
    def _parameter(raw, message):
        text = str(raw if raw is not None else "").strip()
        if not text or len(text) > 64:
            raise ValidationError(message)
        try:
            value = Decimal(text)
        except InvalidOperation as exc:
            raise ValidationError(message) from exc
        if not value.is_finite() or value <= 0 or len(value.as_tuple().digits) > 30:
            raise ValidationError(message)
        return value

    def _validate_archive(self, data):
        if len(data) > self.settings.max_excel_upload_mb * 1024 * 1024:
            raise ValidationError("حجم ملف Excel يتجاوز الحد المسموح.")
        try:
            with zipfile.ZipFile(BytesIO(data)) as archive:
                infos = archive.infolist()
                names = {item.filename for item in infos}
                if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                if len(infos) > self.settings.max_zip_entries:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                limit = self.settings.max_uncompressed_import_mb * 1024 * 1024
                if sum(item.file_size for item in infos) > limit:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
                for item in infos:
                    path = PurePosixPath(item.filename)
                    if item.flag_bits & 1 or path.is_absolute() or ".." in path.parts:
                        raise ValidationError("الملف تالف أو غير مدعوم.")
                    if "\\" in item.filename:
                        raise ValidationError("الملف تالف أو غير مدعوم.")
                bad = archive.testzip()
                if bad:
                    raise ValidationError("الملف تالف أو غير مدعوم.")
        except (zipfile.BadZipFile, OSError) as exc:
            raise ValidationError("الملف تالف أو غير مدعوم.") from exc

    def _transform_sync(self, data, rate, shipping_cost):
        formula_book = value_book = None
        try:
            formula_book = load_workbook(BytesIO(data), data_only=False, keep_links=True)
            value_book = load_workbook(BytesIO(data), data_only=True, keep_links=True)
            processed = 0
            for sheet in formula_book.worksheets:
                value_sheet = value_book[sheet.title]
                header = self._find_header(sheet)
                if header is None:
                    continue
                row_number, sources = header
                result_columns = self._result_columns(sheet, row_number)
                self._calculate_rows(
                    sheet, value_sheet, row_number, sources, result_columns, rate, shipping_cost
                )
                processed += 1
            if not processed:
                raise ValidationError("لم يتم العثور على أعمدة PRICE وPCS وCBM في الملف.")
            output = BytesIO()
            formula_book.save(output)
            return output.getvalue()
        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError("الملف تالف أو غير مدعوم.") from exc
        finally:
            if formula_book is not None:
                formula_book.close()
            if value_book is not None:
                value_book.close()

    def _find_header(self, sheet):
        complete = []
        partial = []
        for row in range(1, min(50, sheet.max_row) + 1):
            matches = {key: [] for key in ALIASES}
            for cell in sheet[row]:
                normalized = normalize_header(cell.value)
                for key, aliases in ALIASES.items():
                    if normalized in aliases:
                        matches[key].append(cell.column)
            present = [key for key, columns in matches.items() if columns]
            if present:
                partial.append((row, matches))
            if len(present) == 3:
                for key, columns in matches.items():
                    if len(columns) != 1:
                        raise ValidationError(
                            f"تعذر تحديد عمود {key} بشكل فريد في الورقة {sheet.title}."
                        )
                columns = {key: values[0] for key, values in matches.items()}
                if len(set(columns.values())) != 3:
                    raise ValidationError(f"تعذر تحديد أعمدة البيانات في الورقة {sheet.title}.")
                complete.append((row, columns))
        if len(complete) > 1:
            raise ValidationError(f"تم العثور على أكثر من صف عناوين في الورقة {sheet.title}.")
        if complete:
            return complete[0]
        if partial:
            best_row, matches = max(
                partial, key=lambda item: sum(bool(x) for x in item[1].values())
            )
            del best_row
            missing = [key for key, columns in matches.items() if not columns]
            if len(missing) == 1:
                raise ValidationError(f"الورقة {sheet.title}: العمود {missing[0]} مفقود.")
        return None

    def _result_columns(self, sheet, row):
        normalized_results = {normalize_header(value): value for value in RESULT_HEADERS}
        found = {value: [] for value in RESULT_HEADERS}
        last_real = 0
        for cell in sheet[row]:
            if cell.value is not None and str(cell.value).strip():
                last_real = cell.column
                result = normalized_results.get(normalize_header(cell.value))
                if result:
                    found[result].append(cell.column)
        for result, columns in found.items():
            if len(columns) > 1:
                raise ValidationError(f"عنوان النتيجة {result} مكرر في الورقة {sheet.title}.")
        result_columns = {}
        template = sheet.cell(row, last_real) if last_real else None
        for result in RESULT_HEADERS:
            if found[result]:
                result_columns[result] = found[result][0]
                continue
            last_real += 1
            cell = sheet.cell(row, last_real, result)
            if template:
                self._copy_style(template, cell)
            sheet.column_dimensions[get_column_letter(last_real)].width = 18
            result_columns[result] = last_real
        return result_columns

    def _calculate_rows(self, sheet, values, header_row, sources, results, rate, shipping_cost):
        for row in range(header_row + 1, sheet.max_row + 1):
            raw = {key: values.cell(row, column).value for key, column in sources.items()}
            formula = {key: sheet.cell(row, column).value for key, column in sources.items()}
            if not any(value is not None and str(value).strip() for value in formula.values()):
                continue
            row_text = " ".join(
                normalize_header(sheet.cell(row, column).value)
                for column in range(1, min(sheet.max_column, 12) + 1)
            )
            if re.search(r"\b(TOTAL|GRAND TOTAL|SUBTOTAL)\b", row_text):
                continue
            parsed = {}
            for key in ("PRICE", "PCS", "CBM"):
                value = raw[key]
                if isinstance(formula[key], str) and formula[key].startswith("=") and value is None:
                    raise ValidationError(
                        f"الورقة {sheet.title}، الصف {row}: قيمة {key} المحسوبة غير متاحة."
                    )
                parsed[key] = self._cell_decimal(value, key, sheet.title, row)
            price, pcs, cbm = parsed["PRICE"], parsed["PCS"], parsed["CBM"]
            price_ouguiya = price * rate
            office_fee = price_ouguiya * OFFICE_RATE
            shipping = cbm * shipping_cost / pcs
            calculated = (
                price_ouguiya,
                office_fee,
                shipping,
                price_ouguiya + office_fee + shipping,
            )
            adjacent = sheet.cell(row, max(sources.values()))
            for heading, value in zip(RESULT_HEADERS, calculated, strict=True):
                target = sheet.cell(row, results[heading])
                if not target.has_style:
                    self._copy_style(adjacent, target)
                rounded = value.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
                target.value = float(rounded)
                target.number_format = "0.######"

    @staticmethod
    def _cell_decimal(raw, name, sheet, row):
        if isinstance(raw, bool) or raw is None:
            raise ValidationError(f"الورقة {sheet}، الصف {row}: قيمة {name} غير صالحة.")
        try:
            if isinstance(raw, float) and not math.isfinite(raw):
                raise InvalidOperation
            value = Decimal(str(raw).strip().replace(",", ""))
        except (InvalidOperation, ValueError):
            raise ValidationError(f"الورقة {sheet}، الصف {row}: قيمة {name} غير صالحة.") from None
        if not value.is_finite() or value < 0:
            raise ValidationError(f"الورقة {sheet}، الصف {row}: قيمة {name} غير صالحة.")
        if name == "PCS" and value == 0:
            raise ValidationError(f"الورقة {sheet}، الصف {row}: قيمة PCS تساوي صفرًا.")
        return value

    @staticmethod
    def _copy_style(source, target):
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        target.number_format = source.number_format
