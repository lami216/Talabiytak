from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.services.errors import ValidationError
from app.services.excel_pricing import RESULT_HEADERS


def workbook_bytes(headers=("PRICE", "PCS", "CBM"), header_row=1, existing=False):
    book = Workbook()
    sheet = book.active
    sheet.title = "منتجات"
    for column, heading in enumerate(headers, 1):
        sheet.cell(header_row, column, heading)
    if existing:
        for offset, heading in enumerate(RESULT_HEADERS, len(headers) + 1):
            sheet.cell(header_row, offset, heading)
    positions = {}
    for index, heading in enumerate(headers, 1):
        if "PRICE" in heading:
            positions["PRICE"] = index
        if "PCS" in heading and "T.T" not in heading:
            positions["PCS"] = index
        if "CBM" in heading and "T.T" not in heading:
            positions["CBM"] = index
    sheet.cell(header_row + 1, positions["PRICE"], 2.7)
    sheet.cell(header_row + 1, positions["PCS"], 40)
    sheet.cell(header_row + 1, positions["CBM"], 0.091124)
    output = BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


def results(data):
    book = load_workbook(BytesIO(data), data_only=True)
    sheet = book["منتجات"]
    headings = {cell.value: cell.column for cell in sheet[1] if cell.value}
    values = tuple(sheet.cell(2, headings[name]).value for name in RESULT_HEADERS)
    book.close()
    return values


@pytest.mark.asyncio
async def test_pricing_uses_request_parameters_and_exact_calculation(setup):
    _, app, *_ = setup
    source = workbook_bytes()
    first = await app.state.excel_pricing.transform(source, "63", "60000")
    second = await app.state.excel_pricing.transform(source, "65", "70000")
    assert results(first) == pytest.approx((170.1, 5.103, 136.686, 311.889))
    assert results(first) != results(second)


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("headers", "header_row"),
    [
        (("X", "PRICE", "Y", "CBM", "Z", "PCS"), 1),
        (("PCS/CTN", "Extra", "CBM PER CTN", "UNIT PRICE RMB"), 3),
        (("CBM", "PCS", "PRICE"), 8),
    ],
)
async def test_columns_and_header_rows_can_move(setup, headers, header_row):
    _, app, *_ = setup
    output = await app.state.excel_pricing.transform(
        workbook_bytes(headers, header_row), "10", "100"
    )
    book = load_workbook(BytesIO(output), data_only=True)
    sheet = book.active
    assert [
        sheet.cell(header_row, sheet.max_column - 3 + index).value for index in range(4)
    ] == list(RESULT_HEADERS)
    book.close()


@pytest.mark.asyncio
async def test_forbidden_total_aliases_are_not_selected(setup):
    _, app, *_ = setup
    data = workbook_bytes(("AMOUNT", "T.T PCS", "T.T CBM", "PRICE", "PCS", "CBM"))
    output = await app.state.excel_pricing.transform(data, "10", "100")
    assert results(output)[0] == pytest.approx(27)


@pytest.mark.asyncio
async def test_existing_results_are_updated_without_duplicates(setup):
    _, app, *_ = setup
    output = await app.state.excel_pricing.transform(workbook_bytes(existing=True), "10", "100")
    book = load_workbook(BytesIO(output))
    headings = [cell.value for cell in book.active[1]]
    assert all(headings.count(name) == 1 for name in RESULT_HEADERS)
    book.close()


@pytest.mark.asyncio
async def test_multiple_sheets_and_incomplete_sheet(setup):
    _, app, *_ = setup
    source = workbook_bytes()
    book = load_workbook(BytesIO(source))
    second = book.create_sheet("ثانية")
    second.append(["CBM", "PRICE", "PCS"])
    second.append([1, 2, 4])
    notes = book.create_sheet("ملاحظات")
    notes["A1"] = "unchanged"
    output = BytesIO()
    book.save(output)
    book.close()
    transformed = await app.state.excel_pricing.transform(output.getvalue(), "10", "100")
    result_book = load_workbook(BytesIO(transformed))
    assert result_book["ملاحظات"]["A1"].value == "unchanged"
    assert result_book["ثانية"]["D1"].value == RESULT_HEADERS[0]
    result_book.close()

    incomplete = Workbook()
    incomplete.active.append(["PRICE", "CBM"])
    buffer = BytesIO()
    incomplete.save(buffer)
    with pytest.raises(ValidationError, match="PCS"):
        await app.state.excel_pricing.transform(buffer.getvalue(), "10", "100")


def test_pricing_routes_auth_csrf_validation_and_download(auth):
    client, _, _, _, token, _ = auth
    client.cookies.clear()
    assert client.get("/pricing", follow_redirects=False).status_code == 303
    assert client.post("/pricing", follow_redirects=False).status_code == 303
    client.post("/login", data={"username": "admin", "password": "strong-password"})
    bad = client.post(
        "/pricing",
        files={"file": ("bad.xlsx", b"bad", "application/octet-stream")},
        data={"csrf_token": "bad", "rmb_rate": "1", "shipping_cost_per_cbm": "1"},
    )
    assert bad.status_code == 400
    token = client.app.state.security.load(
        client.cookies[client.app.state.settings.session_cookie_name]
    )["csrf"]
    response = client.post(
        "/pricing",
        files={"file": ("ملف.xlsx", workbook_bytes(), "application/octet-stream")},
        data={"csrf_token": token, "rmb_rate": "63", "shipping_cost_per_cbm": "60000"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    load_workbook(BytesIO(response.content)).close()


@pytest.mark.asyncio
@pytest.mark.parametrize(("rate", "shipping"), [("", "1"), ("0", "1"), ("-1", "1"), ("1", "NaN")])
async def test_parameters_must_be_positive_finite(setup, rate, shipping):
    _, app, *_ = setup
    with pytest.raises(ValidationError):
        await app.state.excel_pricing.transform(workbook_bytes(), rate, shipping)
