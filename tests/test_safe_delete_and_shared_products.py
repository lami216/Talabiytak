from io import BytesIO

import httpx
from bson import ObjectId
from openpyxl import load_workbook
from PIL import Image

from app.models import ImageAsset, ImageStatus, ImportedImage, Product
from app.utils.objectid import new_id, to_object_id


def asset(file_id="shared-file"):
    return ImageAsset(
        file_id,
        f"/{file_id}.png",
        f"https://ik.imagekit.io/test/{file_id}.png",
        None,
        file_id,
        "image/png",
        20,
        20,
    )


async def add_import(app, image):
    batch = await app.state.repositories.imports.create("batch.xlsx")
    image.import_id = batch.id
    await app.state.repositories.images.create(image)
    return batch, image


def test_delete_unique_and_shared_imported_images(auth):
    client, app, fake, _, token, database = auth
    unique = ImportedImage(
        new_id(),
        "",
        1,
        "xl/media/1.png",
        status=ImageStatus.unnamed.value,
        image_asset=asset("unique"),
    )
    batch, unique = client.portal.call(add_import, app, unique)
    response = client.post(
        f"/imports/images/{unique.id}/delete",
        data={"csrf_token": token, "return_status": "unnamed", "return_page": "1"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == f"/imports/{batch.id}?status=unnamed&page=1"
    assert fake.files.deleted == [{"file_id": "unique"}]
    assert database.raw.imported_images.find_one({"_id": to_object_id(unique.id)}) is None

    shared = asset("shared")
    first = ImportedImage(
        new_id(), "", 1, "a.png", status=ImageStatus.unnamed.value, image_asset=shared
    )
    second = ImportedImage(
        new_id(), "", 2, "b.png", status=ImageStatus.duplicate.value, image_asset=shared
    )
    batch, first = client.portal.call(add_import, app, first)
    second.import_id = batch.id
    client.portal.call(app.state.repositories.images.create, second)
    fake.files.deleted.clear()
    client.post(
        f"/imports/images/{first.id}/delete",
        data={"csrf_token": token, "return_status": "unnamed", "return_page": "1"},
        follow_redirects=False,
    )
    assert fake.files.deleted == []
    assert database.raw.imported_images.find_one({"_id": to_object_id(second.id)}) is not None
    client.post(
        f"/imports/images/{second.id}/delete",
        data={"csrf_token": token, "return_status": "duplicate", "return_page": "1"},
        follow_redirects=False,
    )
    assert fake.files.deleted == [{"file_id": "shared"}]


def test_delete_duplicate_shared_with_product_and_saved_protection(auth):
    client, app, fake, _, token, database = auth
    shared = asset("product-file")
    product = Product(new_id(), "قفل 50", "قفل 50", shared)
    client.portal.call(app.state.repositories.products.create, product)
    duplicate = ImportedImage(
        new_id(), "", 1, "dup.png", status=ImageStatus.duplicate.value, image_asset=shared
    )
    batch, duplicate = client.portal.call(add_import, app, duplicate)
    response = client.post(
        f"/imports/images/{duplicate.id}/delete",
        data={"csrf_token": token, "return_status": "duplicate", "return_page": "1"},
        follow_redirects=False,
    )
    assert response.headers["location"] == f"/imports/{batch.id}?status=duplicate&page=1"
    assert fake.files.deleted == []
    assert database.raw.products.find_one({"_id": to_object_id(product.id)}) is not None

    saved = ImportedImage(
        new_id(),
        "",
        2,
        "saved.png",
        status=ImageStatus.saved_as_product.value,
        image_asset=shared,
        linked_product_id=product.id,
    )
    saved.import_id = batch.id
    client.portal.call(app.state.repositories.images.create, saved)
    html = client.get(f"/imports/{batch.id}?status=saved_as_product").text
    assert "حذف الصورة" not in html
    direct = client.post(
        f"/imports/images/{saved.id}/delete", data={"csrf_token": token}, follow_redirects=False
    )
    assert direct.status_code == 303
    assert database.raw.imported_images.find_one({"_id": to_object_id(saved.id)}) is not None


def test_delete_csrf_and_imagekit_failure(auth):
    client, app, fake, _, token, database = auth
    image = ImportedImage(
        new_id(), "", 1, "x.png", status=ImageStatus.unnamed.value, image_asset=asset("fail")
    )
    _batch, image = client.portal.call(add_import, app, image)
    assert client.post(f"/imports/images/{image.id}/delete", data={}).status_code == 400
    fake.files.fail_delete = True
    response = client.post(
        f"/imports/images/{image.id}/delete", data={"csrf_token": token}, follow_redirects=False
    )
    assert response.status_code == 303
    assert database.raw.imported_images.find_one({"_id": to_object_id(image.id)}) is not None


def test_shared_product_creation_search_order_excel_and_delete(auth):
    client, app, fake, _, token, database = auth
    original = Product(new_id(), "قفل 50", "قفل 50", asset("locks"))
    client.portal.call(app.state.repositories.products.create, original)
    before_uploads = len(fake.files.uploads)
    response = client.post(
        f"/products/{original.id}/create-with-same-image",
        data={"csrf_token": token, "name": "قفل 60"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    created_id = response.headers["location"].split("/")[2]
    created = client.portal.call(app.state.repositories.products.get, created_id)
    assert created.id != original.id
    assert created.primary_image.file_id == original.primary_image.file_id
    assert len(fake.files.uploads) == before_uploads
    search = client.get("/orders/product-search?q=قفل").json()["items"]
    assert {item["id"] for item in search} == {original.id, created.id}
    assert len({item["image_url"] for item in search}) == 1

    order = client.post(
        "/orders/new",
        data={
            "csrf_token": token,
            "title": "مشتركة",
            "product_id": [original.id, created.id],
            "quantity": ["2", "5"],
        },
        follow_redirects=False,
    )
    assert order.status_code == 303
    saved = database.raw.orders.find_one({"title": "مشتركة"})
    assert [(i["product_id"], i["quantity"]) for i in saved["items"]] == [
        (ObjectId(original.id), 2),
        (ObjectId(created.id), 5),
    ]

    image_bytes = BytesIO()
    Image.new("RGB", (20, 20), "blue").save(image_bytes, "PNG")
    app.state.excel_export.transport = httpx.MockTransport(
        lambda _r: httpx.Response(
            200, content=image_bytes.getvalue(), headers={"content-type": "image/png"}
        )
    )
    xlsx = client.get(order.headers["location"] + "/download")
    wb = load_workbook(BytesIO(xlsx.content))
    ws = wb["الطلبية"]
    assert [ws.cell(4, 2).value, ws.cell(5, 2).value] == ["قفل 50", "قفل 60"]
    wb.close()

    client.post(
        f"/products/{original.id}/delete",
        data={"csrf_token": token, "confirm": "delete"},
        follow_redirects=False,
    )
    # Active order protects deletion, so remove order then test reference-safe storage deletion.
    database.raw.orders.delete_many({})
    client.post(
        f"/products/{original.id}/delete",
        data={"csrf_token": token, "confirm": "delete"},
        follow_redirects=False,
    )
    assert fake.files.deleted == []
    assert database.raw.products.find_one({"_id": ObjectId(created.id)}) is not None
    client.post(
        f"/products/{created.id}/delete",
        data={"csrf_token": token, "confirm": "delete"},
        follow_redirects=False,
    )
    assert fake.files.deleted == [{"file_id": "locks"}]


def test_batch_and_product_form_contracts():
    batch = (
        open("app/templates/batch.html", encoding="utf-8").read()
        + open("app/templates/partials/batch_content.html", encoding="utf-8").read()
    )
    assert "حذف الصور غير المستخدمة من ImageKit" not in batch
    assert "تجاهل" not in batch
    assert "متجاهلة" not in batch
    assert "/ignore" not in batch
    assert "حذف الصورة" in batch
    assert "/delete" in batch
    assert "هل تريد حذف هذه الصورة؟ لا يمكن التراجع عن هذه العملية." in batch
    product_form = open("app/templates/product_form.html", encoding="utf-8").read()
    assert "إضافة منتج آخر بنفس الصورة" in product_form
    assert "إنشاء منتج بنفس الصورة" in product_form


def test_shared_product_ajax_json_contract(auth):
    client, app, fake, _, token, _database = auth
    original = Product(new_id(), "قفل 50", "قفل 50", asset("ajax-locks"))
    client.portal.call(app.state.repositories.products.create, original)
    before_uploads = len(fake.files.uploads)
    response = client.post(
        f"/products/{original.id}/create-with-same-image",
        data={"csrf_token": token, "name": "قفل 60"},
        headers={"X-Requested-With": "fetch"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["product"]["name"] == "قفل 60"
    assert payload["product"]["edit_url"].endswith("/edit")
    created = client.portal.call(app.state.repositories.products.get, payload["product"]["id"])
    assert created.id != original.id
    assert created.primary_image.file_id == original.primary_image.file_id
    assert len(fake.files.uploads) == before_uploads


def test_batch_partial_get_and_delete_ajax_contract(auth):
    client, app, fake, _, token, _database = auth
    image = ImportedImage(
        new_id(),
        "",
        1,
        "one.png",
        status=ImageStatus.unnamed.value,
        image_asset=asset("ajax-delete"),
    )
    batch, image = client.portal.call(add_import, app, image)
    full = client.get(f"/imports/{batch.id}?status=invalid&page=-5")
    assert full.status_code == 200
    assert "<!doctype html>" in full.text
    partial = client.get(
        f"/imports/{batch.id}?status=invalid&page=-5", headers={"X-Requested-With": "fetch"}
    )
    assert partial.status_code == 200
    payload = partial.json()
    assert payload["ok"] is True
    assert payload["status"] == "all"
    assert payload["page"] == 1
    assert "batch-filters" in payload["html"]
    deleted = client.post(
        f"/imports/images/{image.id}/delete",
        data={"csrf_token": token, "return_status": "unnamed", "return_page": "1"},
        headers={"X-Requested-With": "fetch"},
    )
    assert deleted.status_code == 200
    assert deleted.json()["ok"] is True
    assert deleted.json()["result"]["storage_deleted"] is True
    assert fake.files.deleted == [{"file_id": "ajax-delete"}]
