from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def tiny_png(tmp_path):
    path = tmp_path / "tiny.png"
    with Image.new("RGB", (2, 2), "red") as img:
        img.save(path)
    return path


def pricing_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "LMT"
    headers = ["CODE", "NAME", "PRICE", "PCS", "CBM", "F", "G", "H", "I", "J", "K", "L"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    rows = [
        ("LMT-1", "full", 10, 20, 0.02),
        ("LMT-8-3", "missing price", None, 50, 0.039204),
        ("LMT-8-4", "missing price", "", 50, 0.039204),
        ("LMT-0", "zero pcs", 10, 0, 0.039204),
        ("LMT-M", "missing both", None, None, 0.039204),
        ("Total", "", None, None, None),
    ]
    for r, row in enumerate(rows, 4):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    drawing = XLImage(str(tiny_png(tmp_path)))
    drawing.width = 130
    ws.add_image(drawing, "M4")
    out = BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def test_partial_pricing_keeps_images_and_uses_safe_columns(auth, tmp_path):
    client, *_rest, token, _database = auth
    response = client.post(
        "/pricing",
        data={"csrf_token": token, "rmb_rate": "63", "shipping_cost_per_cbm": "60000"},
        files={"file": ("LMT-test.xlsx", pricing_workbook(tmp_path), MIME)},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == MIME
    wb = load_workbook(BytesIO(response.content))
    ws = wb["LMT"]
    assert len(ws._images) == 1
    assert ws._images[0].anchor._from.col == 12
    assert [ws.cell(3, c).value for c in range(15, 19)] == [
        "السعر بالأوقية",
        "نسبة المكتب",
        "الشحن",
        "سعر الطياح",
    ]
    assert ws.cell(4, 15).value == 630
    assert ws.cell(4, 16).value == 18.9
    assert ws.cell(4, 17).value == 60
    assert ws.cell(4, 18).value == 708.9
    for row in (5, 6):
        assert ws.cell(row, 15).value is None
        assert ws.cell(row, 16).value is None
        assert ws.cell(row, 17).value == pytest.approx(47.0448)
        assert ws.cell(row, 18).value == pytest.approx(47.0448)
    assert ws.cell(7, 17).value is None
    assert ws.cell(7, 18).value == pytest.approx(648.9)
    assert ws.cell(8, 15).value is None
    assert ws.cell(8, 18).value is None
    assert "#DIV/0!" not in [cell.value for row in ws.iter_rows() for cell in row]
    with ZipFile(BytesIO(response.content)) as archive:
        assert any(name.startswith("xl/media/") for name in archive.namelist())
    wb.close()
