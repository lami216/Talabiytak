from datetime import timedelta
from io import BytesIO
from zipfile import ZipFile

import httpx
import pytest
from openpyxl import load_workbook
from PIL import Image

from app.database.mongo import REQUIRED_INDEXES
from app.models import ImageAsset, Order, OrderItem, Product, now
from app.services.excel_export import ExcelExportService
from app.utils.objectid import new_id


async def product(repository, name):
    item = Product(
        new_id(),
        name,
        name,
        ImageAsset("file", f"/{name}.png", "ignored", None, name, "image/png", 20, 20),
    )
    await repository.create(item)
    return item


@pytest.mark.asyncio
async def test_order_repository_lifecycle_and_indexes(setup):
    _, app, _, _, database = setup
    first = await product(app.state.repositories.products, "أول")
    second = await product(app.state.repositories.products, "ثان")
    order = await app.state.orders.create("  طلبية الساحة  ", [first.id, second.id], [2, 7])
    document = database.raw.orders.find_one()
    assert order.title == "طلبية الساحة"
    assert [item.position for item in order.items] == [1, 2]
    assert document["items"][0]["product_id"] != first.id
    assert not any(isinstance(value, bytes) for value in document.values())
    original_expiry = document["expires_at"]
    await app.state.orders.update(order.id, "معدلة", [second.id], [9])
    updated = await app.state.repositories.orders.get_active(order.id)
    assert updated.expires_at == original_expiry
    assert updated.items[0].product_name == "ثان"
    assert "orders_expires_at_ttl" in REQUIRED_INDEXES["orders"]


@pytest.mark.asyncio
async def test_expired_orders_are_hidden(setup):
    _, app, _, _, database = setup
    item = await product(app.state.repositories.products, "قديم")
    order = await app.state.orders.create("منتهية", [item.id], [1])
    database.raw.orders.update_one(
        {"_id": document_id(database, order.id)},
        {"$set": {"expires_at": now() - timedelta(days=1)}},
    )
    assert await app.state.repositories.orders.get_active(order.id) is None
    assert not await app.state.repositories.orders.list_active()


def document_id(database, value):
    return database.raw.orders.find_one({"title": {"$exists": True}})["_id"]


def test_order_routes_excel_and_protection(auth):
    client, app, _, _, token, database = auth
    image = BytesIO()
    with Image.new("RGB", (20, 10), "red") as picture:
        picture.save(image, "PNG")

    def delivery(_request):
        return httpx.Response(200, content=image.getvalue(), headers={"content-type": "image/png"})

    app.state.excel_export.transport = httpx.MockTransport(delivery)
    repository = app.state.repositories.products
    first = client.portal.call(product, repository, "حليب")
    response = client.post(
        "/orders/new",
        data={"csrf_token": token, "title": "طلبية عربية", "product_id": first.id, "quantity": "4"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    download_url = response.headers["location"] + "/download"
    response = client.get(download_url)
    assert response.status_code == 200
    assert response.content.startswith(b"PK")
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["الطلبية"]
    assert sheet.sheet_view.rightToLeft is False
    assert (
        sheet["A1"].value,
        sheet["A3"].value,
        sheet["B3"].value,
        sheet["C3"].value,
        sheet["B4"].value,
        sheet["C4"].value,
    ) == (
        "طلبية عربية",
        "الصورة",
        "اسم المنتج",
        "الكمية",
        "حليب",
        4,
    )
    assert sheet.column_dimensions["A"].width == pytest.approx(9.5)
    assert sheet.column_dimensions["B"].width == pytest.approx(25)
    assert sheet.column_dimensions["C"].width == pytest.approx(9)
    assert sheet.row_dimensions[1].height == 26
    assert sheet.row_dimensions[3].height == 22
    assert sheet.row_dimensions[4].height == 54
    assert sheet["A1"].font.sz == 16
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.color.rgb == "00000000"
    for header in sheet[3]:
        assert header.font.sz == 13
        assert header.font.bold is True
        assert header.font.color.rgb == "00000000"
    assert len(sheet._images) == 1
    workbook.close()
    assert client.get("/orders/product-search?q=x").status_code == 200
    client.cookies.clear()
    assert client.get(download_url, follow_redirects=False).status_code == 303


def test_excel_images_crop_to_fill_and_anchor_to_product_rows():
    sizes = ((200, 100), (100, 200), (80, 80))
    pictures = []
    for size in sizes:
        output = BytesIO()
        with Image.new("RGB", size, "green") as source:
            if size[0] > size[1]:
                source.paste("red", (0, 0, 50, size[1]))
                source.paste("blue", (150, 0, size[0], size[1]))
            elif size[1] > size[0]:
                source.paste("red", (0, 0, size[0], 50))
                source.paste("blue", (0, 150, size[0], size[1]))
            source.save(output, "PNG")
        pictures.append(output.getvalue())
    order = Order(
        "order-id",
        "طلبية الصور",
        [OrderItem(f"product-{index}", f"منتج {index}", index, index) for index in range(1, 4)],
    )

    data = ExcelExportService._workbook(order, pictures)

    workbook = load_workbook(BytesIO(data))
    sheet = workbook["الطلبية"]
    assert sheet.sheet_view.rightToLeft is False
    assert sheet.column_dimensions["A"].width == pytest.approx(9.5)
    assert sheet.column_dimensions["B"].width == pytest.approx(25)
    assert sheet.column_dimensions["C"].width == pytest.approx(9)
    assert len(sheet._images) == len(order.items)
    for row, embedded in enumerate(sheet._images, 4):
        assert (embedded.width, embedded.height) == (72, 72)
        assert embedded.anchor._from.col == 0
        assert embedded.anchor._from.row == row - 1
        assert embedded.anchor._from.colOff == 0
        assert embedded.anchor._from.rowOff == 0
        assert embedded.anchor.ext.cx == 72 * 9525
        assert embedded.anchor.ext.cy == 72 * 9525
        assert sheet.row_dimensions[row].height == 54
        assert sheet.cell(row, 2).value == f"منتج {row - 3}"
        assert sheet.cell(row, 2).font.sz == 14
        assert sheet.cell(row, 2).font.bold is True
        assert sheet.cell(row, 2).font.color.rgb == "00000000"
        assert sheet.cell(row, 2).alignment.horizontal == "center"
        assert sheet.cell(row, 2).alignment.vertical == "center"
        assert sheet.cell(row, 2).alignment.wrap_text is True
        assert sheet.cell(row, 2).alignment.readingOrder == 2
        assert sheet.cell(row, 3).value == row - 3
        assert isinstance(sheet.cell(row, 3).value, int)
        assert sheet.cell(row, 3).font.sz == 13
        assert sheet.cell(row, 3).font.bold is True
        assert sheet.cell(row, 3).font.color.rgb == "00000000"
        assert sheet.cell(row, 3).alignment.horizontal == "center"
        assert sheet.cell(row, 3).alignment.vertical == "center"
    workbook.close()

    with ZipFile(BytesIO(data)) as archive:
        media = sorted(name for name in archive.namelist() if name.startswith("xl/media/"))
        assert len(media) == len(sizes)
        for name in media:
            with Image.open(BytesIO(archive.read(name))) as embedded:
                assert embedded.size == (72, 72)
                # Colored outer bands are removed by a centered cover crop, not squeezed in.
                edge_midpoints = ((0, 36), (71, 36), (36, 0), (36, 71))
                for pixel in (embedded.getpixel(point) for point in edge_midpoints):
                    red, green, blue = pixel
                    assert green > red
                    assert green > blue
